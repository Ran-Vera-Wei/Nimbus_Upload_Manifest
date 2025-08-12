import io
import re
import numpy as np
import pandas as pd
import msoffcrypto
import streamlit as st
from openpyxl import load_workbook

# -----------------------
# Streamlit UI
# -----------------------
st.set_page_config(page_title="Excel Converter (Decrypt + Transform)", page_icon="📄")
st.title("📄 Excel Converter (Password Remove + Transform)")

st.markdown("""
**What this app does**
1. Decrypts an uploaded **.xlsx** using the password you provide.  
2. Processes **hawb** with two header rows (row 2 = real header, data starts row 3).  
3. Applies your requested transformations to **hawb** and updates **mawb!L2**.  
4. Returns a clean, transformed file for download.
""")

default_pw = "_S8&Dwy2&U"
password = st.text_input("File password", value=default_pw, type="password")
uploaded = st.file_uploader("Upload password-protected .xlsx", type=["xlsx"])

run_btn = st.button("Convert")

# -----------------------
# Helpers
# -----------------------
def decrypt_xlsx(file_bytes: bytes, password: str) -> bytes:
    """Return decrypted xlsx as bytes."""
    fin = io.BytesIO(file_bytes)
    fout = io.BytesIO()
    office_file = msoffcrypto.OfficeFile(fin)
    office_file.load_key(password=password)
    office_file.decrypt(fout)
    return fout.getvalue()

def normalize_map(cols):
    """Map normalized (lower, trim) -> original column names."""
    mapping = {}
    for c in cols:
        key = str(c).strip().lower()
        # If duplicate normalized names exist, keep first occurrence
        mapping.setdefault(key, c)
    return mapping

def find_col(cols_map, name):
    """Find original column by case-insensitive, trimmed match."""
    return cols_map.get(str(name).strip().lower(), None)

def keep_half_if_over(s, limit):
    """If string length > limit, keep only first half; otherwise return original."""
    if pd.isna(s):
        return s
    s = str(s)
    if len(s) > limit:
        # floor half
        return s[: len(s) // 2]
    return s

def enforce_zip_6_digits(x):
    """If not exactly 6 digits, set to '123456'."""
    if pd.isna(x):
        return "123456"
    s = str(x).strip()
    return s if re.fullmatch(r"\d{6}", s) else "123456"

# -----------------------
# Main
# -----------------------
if run_btn:
    if not uploaded:
        st.warning("Please upload a .xlsx file first.")
        st.stop()
    if not password:
        st.warning("Please enter the password.")
        st.stop()

    try:
        with st.spinner("Decrypting..."):
            decrypted = decrypt_xlsx(uploaded.read(), password)

        # Load workbook once (for mawb edit and to preserve other sheets)
        wb = load_workbook(io.BytesIO(decrypted))

        # ---- Update mawb!L2 (under consignee_id_number) ----
        if "mawb" in wb.sheetnames:
            ws = wb["mawb"]
            # Set L2 no matter what (your requirement)
            ws["L2"].value = "2567704"
        else:
            st.warning("Sheet 'mawb' not found. Skipping mawb update.")

        # ---- Process hawb sheet using 2-row header logic ----
        if "hawb" not in wb.sheetnames:
            st.error("Sheet 'hawb' not found in workbook.")
            st.stop()

        # Read hawb as raw (no header) to reconstruct headers correctly
        raw_hawb = pd.read_excel(io.BytesIO(decrypted), sheet_name="hawb", header=None, dtype=str)

        if raw_hawb.shape[0] < 2:
            st.error("The 'hawb' sheet does not contain at least two header rows.")
            st.stop()

        # Row index 1 (2nd row) is the real header; data starts at row index 2 (3rd row)
        new_cols = raw_hawb.iloc[1].astype(str).tolist()
        hawb_df = raw_hawb.iloc[2:].copy()
        hawb_df.columns = new_cols
        hawb_df.reset_index(drop=True, inplace=True)

        # Build a map of normalized header -> original header (to handle spaces/case)
        colmap = normalize_map(hawb_df.columns)

        # 1) Half-length trimming rules
        # manufacture_name: if len > 100 -> keep first half
        col_manu_name = find_col(colmap, "manufacture_name")
        if col_manu_name in hawb_df.columns:
            hawb_df[col_manu_name] = hawb_df[col_manu_name].apply(lambda x: keep_half_if_over(x, 100))

        # manufacture_address: if len > 225 -> keep first half
        col_manu_addr = find_col(colmap, "manufacture_address")
        if col_manu_addr in hawb_df.columns:
            hawb_df[col_manu_addr] = hawb_df[col_manu_addr].apply(lambda x: keep_half_if_over(x, 225))

        # manufacture_state: if len > 8 -> keep first half
        col_manu_state = find_col(colmap, "manufacture_state")
        if col_manu_state in hawb_df.columns:
            hawb_df[col_manu_state] = hawb_df[col_manu_state].apply(lambda x: keep_half_if_over(x, 8))

        # 2) Set all country_of_origin and manufacture_country to "CN"
        for cname in ["country_of_origin", "manufacture_country"]:
            col = find_col(colmap, cname)
            if col in hawb_df.columns:
                hawb_df[col] = "CN"

        # 3) Zip rule: if value not exactly 6 digits, fill "123456"
        # Note: original may be 'manufacture_zip_code ' with trailing space
        # Try both exact and trimmed lookups
        col_zip_exact = find_col(colmap, "manufacture_zip_code ")
        col_zip_trim = find_col(colmap, "manufacture_zip_code")
        col_zip = col_zip_exact if col_zip_exact in hawb_df.columns else col_zip_trim
        if col_zip in hawb_df.columns:
            hawb_df[col_zip] = hawb_df[col_zip].apply(enforce_zip_6_digits)

        # 4) Drop recipient_state entirely
        col_recipient_state = find_col(colmap, "recipient_state")
        if col_recipient_state in hawb_df.columns:
            hawb_df.drop(columns=[col_recipient_state], inplace=True)

        # ---- Write back: replace hawb in the existing workbook, keep others (including edited mawb) ----
        # Remove existing hawb so we can write a clean one
        if "hawb" in wb.sheetnames:
            std = wb["hawb"]
            wb.remove(std)
            wb.create_sheet("hawb")  # placeholder to keep desired position (will be overwritten by pandas)

        # Save current wb state to bytes (so pandas can append)
        temp_io = io.BytesIO()
        wb.save(temp_io)

        # Now write the processed hawb into the same workbook
        out_io = io.BytesIO()
        with pd.ExcelWriter(out_io, engine="openpyxl") as writer:
            writer.book = load_workbook(io.BytesIO(temp_io.getvalue()))
            # Align writer.sheets so pandas knows existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

            # Replace hawb sheet cleanly
            if "hawb" in writer.book.sheetnames:
                del writer.book["hawb"]
            hawb_df.to_excel(writer, sheet_name="hawb", index=False)

            writer.save()

        st.success("Conversion complete!")
        st.download_button(
            "Download converted file",
            data=out_io.getvalue(),
            file_name="converted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("What exactly was applied?"):
            st.write("""
- Decrypted the workbook.
- **hawb**:
  - Used 2nd row as headers; data from 3rd row onward.
  - If `manufacture_name` length > 100 → kept first half of the text.
  - If `manufacture_address` length > 225 → kept first half.
  - If `manufacture_state` length > 8 → kept first half.
  - Set `country_of_origin` and `manufacture_country` to `"CN"`.
  - If `manufacture_zip_code` (trailing space tolerated) is not exactly 6 digits → set to `"123456"`.
  - Dropped `recipient_state`.
- **mawb**:
  - Set cell **L2** (under `consignee_id_number`) to `2567704`.
""")

    except msoffcrypto.exceptions.DecryptionError:
        st.error("Decryption failed. Please verify the password.")
    except Exception as e:
        st.exception(e)
